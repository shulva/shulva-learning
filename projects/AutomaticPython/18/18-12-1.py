import openpyxl
import random
import smtplib
from email.header import Header
from email.message import EmailMessage
# pdf-568

mail_list = openpyxl.load_workbook("mail.xlsx")
sheet = mail_list[mail_list.sheetnames[0]]

max_row = sheet.max_row

chores = []
mails = []
for i in range(max_row):
    mails.append(sheet['A'+str(i+1)].value)
    chores.append(sheet['B'+str(i+1)].value)

mails = list(set(mails))
mails.remove(None)

message = EmailMessage()

message['Subject'] = 'Python 邮件自动工作'
message['From'] = "1329068252@qq.com"         #utf-8 会被加密，从而不符合qq邮箱的From Header要求
message['To'] = Header("worker", 'utf-8')     # 接收者

for i in range(len(mails)):
    smtp = smtplib.SMTP('smtp.qq.com', 587)
    smtp.ehlo()
    smtp.starttls()
    smtp.login('1329068252@qq.com', 'yikypytdaarejcig')
    smtp.ehlo()

    random_chore = random.choice(chores)
    message.set_content(str(random_chore))

    smtp.sendmail('1329068252@qq.com', str(random.choice(mails)), message.as_string())
    chores.remove(random_chore)

    smtp.quit()
