#!/usr/bin/env python3
# pdf-369 book-258
import openpyxl
import sys
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string

num = int(sys.argv[1])

workbook = openpyxl.Workbook()
workbook.create_sheet()

font_obj = Font(bold=True)

sheet = workbook['Sheet']

for i in range(1, num+1):
    sheet['A'+str(i+1)].font = font_obj
    sheet['A'+str(i+1)] = i

    sheet[get_column_letter(i+1)+'1'].font = font_obj
    sheet[get_column_letter(i+1)+'1'] = i

for row in range(2, num+2):
    for column in range(2, num+2):
        sheet[get_column_letter(column)+str(row)] = (row-1)*(column-1)

workbook.save('13-14-1.xlsx')
