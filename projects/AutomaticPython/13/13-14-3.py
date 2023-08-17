#!/usr/bin/env python3
# pdf-369 book-258
import openpyxl
import sys
import logging
from openpyxl.utils import get_column_letter, column_index_from_string

logging.basicConfig(level=logging.DEBUG,
                    format=' %(asctime)s - %(levelname)s - %(message)s')

path = sys.argv[1]

workbook = openpyxl.load_workbook(str(path))

first = workbook.sheetnames[0]
sheet = workbook[first]

row_max = sheet.max_row
column_max = sheet.max_column

for row in range(1, row_max+1):
    for column in range(1, column_max+1):
        column_next = get_column_letter(column)
        sheet[column_next+str(row_max+row)
              ].value = sheet[column_next+str(row)].value

for row in range(row_max+1, 2*row_max+1):
    for column in range(1, column_max+1):
        column_next = get_column_letter(column)
        sheet[get_column_letter(row-row_max)+str(column)
              ].value = sheet[column_next+str(row)].value

#logging.debug(row_max)
#logging.debug(column_max)

for row in range(row_max+1, 2*row_max+1):
    for column in range(1, column_max+1):
        column_next = get_column_letter(column)
        sheet[column_next+str(row)].value = ''

if row_max > column_max:
    for row in range(column_max+1, row_max+1):
        for column in range(1, column_max+1):
            column_next = get_column_letter(column)
            sheet[column_next+str(row)].value = ''
elif row_max < column_max:
    for row in range(1, row_max+1):
        for column in range(row_max+1, column_max+1):
            column_next = get_column_letter(column)
            sheet[column_next+str(row)].value = ''

workbook.save(str(path))
