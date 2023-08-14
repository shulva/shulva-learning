#!/usr/bin/env python3
# pdf-369 book-258
import openpyxl
import sys
import logging
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.DEBUG,
                    format=' %(asctime)s - %(levelname)s - %(message)s')

n = int(sys.argv[1])
m = int(sys.argv[2])  # space
file = str(sys.argv[3])

workbook = openpyxl.load_workbook(file)
first = workbook.sheetnames[0]
sheet = workbook[first]

logging.debug("n:"+str(n))

row_max = sheet.max_row
column_max = sheet.max_column

for row in range(row_max, n-1, -1):
    for column in range(1, column_max+1, 1):
        row_next = row + m
        column_next = get_column_letter(column)

#        logging.debug('row_next:'+str(row_next))
#        logging.debug('column_next:'+str(column_next))
        sheet[column_next+str(row_next)
              ].value = sheet[column_next+str(row)].value

for row in range(n, n+m, 1):
    for column in range(1, column_max+1, 1):
        column_next = get_column_letter(column)
        sheet[column_next+str(row)].value = ' '

workbook.save('13-14-1.xlsx')
