#!/usr/bin/env python3
# pdf-369 book-258
import openpyxl
import sys
import logging
from pathlib import Path
from openpyxl.utils import get_column_letter, column_index_from_string

logging.basicConfig(level=logging.DEBUG,
                    format=' %(asctime)s - %(levelname)s - %(message)s')

workbook = openpyxl.load_workbook('13-14-45.xlsx')

first = workbook.sheetnames[0]
sheet = workbook[first]

path = Path(Path.cwd())
file_list = list(path.glob('*.txt'))

for file_index in range(len(file_list)):
    file = open(file_list[file_index])
    #logging.debug(file.name)
    str_list = file.readlines()

    for str_index in range(len(str_list)):
        #logging.debug(len(str_list))
        #logging.debug(str_index)
        sheet[get_column_letter(file_index+1)+str(str_index+1)].value = str_list[str_index]

workbook.save('13-14-45.xlsx')