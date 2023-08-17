#!/usr/bin/env python3
# pdf-369 book-258
import openpyxl
import sys
import logging
from pathlib import Path
from openpyxl.utils import get_column_letter, column_index_from_string

logging.basicConfig(level=logging.DEBUG,
                    format=' %(asctime)s - %(levelname)s - %(message)s')

workbook = openpyxl.load_workbook('13-14-45.xlsx')

first = workbook.sheetnames[0]
sheet = workbook[first]

row_max = sheet.max_row
column_max = sheet.max_column

path = Path(Path.cwd())
file_list = list(path.glob('*.txt'))

for file_index in range(len(file_list)):
    file = open(file_list[file_index], 'w')

    for str_index in range(row_max):
        data = sheet[get_column_letter(file_index+1)+str(str_index+1)].value
        if data != None:
            file.write(str(data))

workbook.save('13-14-45.xlsx')
