#!/usr/bin/env python3
# pdf-430 book-309

import csv
import os
import openpyxl
from pathlib import Path
from openpyxl.utils import get_column_letter

for excelFile in os.listdir('.'):

    # Skip non-xlsx files, load the workbook object.
    if excelFile.endswith(".xlsx"):
        workbook = openpyxl.load_workbook(excelFile)
        path = Path('./'+excelFile)
        name = path.stem

        # Loop through every sheet in the workbook.
        for sheetName in workbook.sheetnames:
            sheet = workbook[sheetName]

            # Create the CSV filename from the Excel filename and sheet title.
            csv_file = open(name+sheetName+'.csv', 'w', newline='')
            # Create the csv.writer object for this CSV file.
            csv_writer = csv.writer(csv_file)

            # Loop through every row in the sheet.
            for rowNum in range(1, sheet.max_row + 1):
                # append each cell to this list
                rowData = []
                # Loop through each cell in the row.
                for colNum in range(1, sheet.max_column + 1):
                    # Append each cell's data to rowData.
                    rowData.append(sheet[get_column_letter(colNum)+str(rowNum)].value)

                # Write the rowData list to the CSV file.
                csv_writer.writerow(rowData)
            csv_file.close()
